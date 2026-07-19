"""3개년 재무제표/재무비율을 그래프 포함 엑셀(.xlsx)로 출력한다.

VBA 매크로 방식(.xlsm) 대신 openpyxl로 네이티브 엑셀 차트를 직접 삽입한다.
매크로가 없으므로 감사법인 보안정책상 매크로 파일이 차단되는 문제가 없고,
받는 사람이 "매크로 사용" 설정 없이 파일을 열자마자 표/그래프를 볼 수 있다.
"""

from __future__ import annotations

import math

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.comments import Comment
from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties, RegularTextRun
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import screening, verification

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
# 조건부 서식(dxf)의 solid 패턴은 fgColor가 아니라 bgColor를 실제 렌더링 색상으로
# 쓴다(OOXML/openpyxl의 잘 알려진 함정) — fgColor만 지정하면 Excel에서 색이 전혀
# 안 보인다. start_color/end_color(=fgColor/bgColor 별칭)를 동일하게 지정해야
# 일반 셀 채우기와 조건부 서식 양쪽 모두에서 올바르게 렌더링된다.
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SCREENING_HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)

# 감사 위험 판단의 핵심 비율만 "메인" 그래프에 태운다(나머지 3개 수익성 비율은 표로만 제공).
MAIN_RATIO_CHART_ITEMS = ["유동비율", "부채비율", "자기자본비율", "ROA", "ROE"]

# 원 항목 그래프는 자산 구성/부채 구성을 각각 스택 바 차트로 분리한다.
# (자산과 부채를 하나의 스택으로 합치면 합계가 자산총계도 부채총계도 아닌
# 의미 없는 숫자가 되므로, 구성비가 각각 명확히 드러나도록 둘로 나눔)
ASSET_STRUCTURE_ITEMS = ["유동자산", "비유동자산"]
LIABILITY_STRUCTURE_ITEMS = ["유동부채", "비유동부채"]


def _style_header_row(ws: Worksheet, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws: Worksheet, n_cols: int, min_width: int = 12) -> None:
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        width = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=min_width,
        )
        ws.column_dimensions[letter].width = max(min_width, width + 2)


def _write_df(ws: Worksheet, df: pd.DataFrame, start_row: int, index_title: str) -> int:
    """DataFrame을 (index_title 포함) 표로 쓰고, 헤더 스타일 적용 후 마지막 행 번호 반환."""
    ws.cell(row=start_row, column=1, value=index_title)
    for j, col in enumerate(df.columns, start=2):
        ws.cell(row=start_row, column=j, value=str(col))
    _style_header_row(ws, start_row, len(df.columns) + 1)

    for i, (idx, row) in enumerate(df.iterrows(), start=start_row + 1):
        ws.cell(row=i, column=1, value=str(idx))
        for j, col in enumerate(df.columns, start=2):
            value = row[col]
            cell = ws.cell(row=i, column=j)
            if value is None or (isinstance(value, float) and math.isnan(value)):
                cell.value = None
            else:
                cell.value = round(float(value), 2)
                cell.number_format = "#,##0.00"

    return start_row + len(df.index)


def _annotate_blank_ratio_cells(
    ws: Worksheet,
    ratio_df: pd.DataFrame,
    table_start: int,
    blank_reasons: dict[str, dict[str, str]] | None,
) -> None:
    """공란 처리된 비율 칸에, 왜 공란인지 설명하는 셀 메모(comment)를 붙인다."""
    if not blank_reasons:
        return
    for name, period_reasons in blank_reasons.items():
        if name not in ratio_df.index:
            continue
        row = table_start + 1 + ratio_df.index.get_loc(name)
        for period, reason in period_reasons.items():
            if period not in ratio_df.columns:
                continue
            col = 2 + ratio_df.columns.get_loc(period)
            ws.cell(row=row, column=col).comment = Comment(reason, "fs-analyze")


def _add_ratio_sheet(
    wb: Workbook,
    ratio_df: pd.DataFrame,
    blank_reasons: dict[str, dict[str, str]] | None = None,
) -> Worksheet:
    ws = wb.create_sheet("재무비율")
    ws.cell(row=1, column=1, value="3개년 재무비율 (단위: %)").font = TITLE_FONT

    table_start = 3
    last_row = _write_df(ws, ratio_df, table_start, "비율명")
    _annotate_blank_ratio_cells(ws, ratio_df, table_start, blank_reasons)
    n_cols = len(ratio_df.columns) + 1

    # 전기대비증감률(%) 컬럼에서 급변동(|증감률|>=20%)을 감사 위험평가 관점에서 강조 표시
    if "전기대비증감률(%)" in ratio_df.columns:
        change_col = ratio_df.columns.get_loc("전기대비증감률(%)") + 2
        col_letter = get_column_letter(change_col)
        cell_range = f"{col_letter}{table_start + 1}:{col_letter}{last_row}"
        ws.conditional_formatting.add(
            cell_range, CellIsRule(operator="greaterThanOrEqual", formula=["20"], fill=WARN_FILL)
        )
        ws.conditional_formatting.add(
            cell_range, CellIsRule(operator="lessThanOrEqual", formula=["-20"], fill=WARN_FILL)
        )

    _autofit(ws, n_cols)
    ws.freeze_panes = ws.cell(row=table_start + 1, column=2)

    # 요청사항: 그래프는 비율표 "옆"에 배치 -> 표 오른쪽에 빈 열 하나를 띄우고 앵커링한다.
    # 비율 그래프는 감사 위험평가의 "메인" 화면 — 원 항목(절대금액) 그래프는 별도
    # 시트("원항목 추이")에 보조/드릴다운 자료로 분리 배치한다.
    chart_anchor_col = get_column_letter(n_cols + 2)
    _add_ratio_chart(ws, ratio_df, table_start, f"{chart_anchor_col}{table_start}")

    return ws


def _add_bs_item_table(ws: Worksheet, item_df: pd.DataFrame, title_row: int) -> int:
    """재무상태표 주요 항목(유동/비유동 자산·부채, 억원) 표를 쓰고 마지막 행 번호를 반환."""
    ws.cell(row=title_row, column=1, value="주요 재무상태표 항목 (단위: 억원)").font = TITLE_FONT
    table_start = title_row + 2
    last_row = _write_df(ws, item_df, table_start, "항목")
    _autofit(ws, len(item_df.columns) + 1)
    return table_start


def _chart_title_with_unit(main_text: str, unit_text: str) -> Title:
    """차트 제목 아래에 단위 표기를 오른쪽 정렬로 붙인 2줄짜리 제목을 만든다.

    세로축 제목으로 단위를 표시하면 눈금 숫자와 겹쳐 보이는 문제가 있어(사용자 피드백),
    플롯 영역 바깥(차트 제목 영역)의 오른쪽 구석에 단위만 별도 줄로 배치한다.
    """
    main_run = RegularTextRun(t=main_text, rPr=CharacterProperties(b=True, sz=1400))
    main_para = Paragraph(pPr=ParagraphProperties(algn="ctr"), r=[main_run])

    unit_run = RegularTextRun(t=unit_text, rPr=CharacterProperties(b=False, sz=1000))
    unit_para = Paragraph(pPr=ParagraphProperties(algn="r"), r=[unit_run])

    return Title(tx=Text(rich=RichText(p=[main_para, unit_para])))


def _configure_axes(chart, height: float = 14, width: float = 26) -> None:
    """openpyxl 라인/바 차트 공통 축 설정.

    openpyxl의 LineChart/BarChart는 x_axis.axPos 기본값이 "l"(왼쪽)이라 카테고리축과
    값축이 같은 위치에 겹쳐 엑셀에서 차트가 깨져 보인다(빈 플레이스홀더처럼 렌더링됨).
    반드시 axPos를 명시적으로 지정해야 한다. 마찬가지로 openpyxl은 axId만 만들고
    delete/tickLblPos를 비워두는데, 그러면 엑셀이 눈금 라벨을 숨겨버리므로
    두 축 모두 명시적으로 "삭제 안 함" + "축 옆에 라벨 표시"로 지정해야 한다.
    """
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    # 차트가 작으면 축 제목이 눈금 숫자/카테고리 라벨과 겹쳐 보인다. 넉넉하게 키운다.
    chart.height, chart.width = height, width
    chart.legend.position = "r"
    chart.legend.overlay = False


def _add_ratio_chart(
    ws: Worksheet, ratio_df: pd.DataFrame, table_start: int, anchor_cell: str
) -> None:
    """감사 위험 판단의 핵심 비율(유동비율/부채비율/자기자본비율/ROA/ROE) 추이를
    라인 차트로 삽입한다. 원 항목 그래프보다 먼저 눈에 들어오는 "메인" 그래프.
    """
    periods_present = [c for c in ratio_df.columns if c != "전기대비증감률(%)"]
    if not periods_present:
        return

    min_col, max_col = 2, 1 + len(periods_present)

    chart = LineChart()
    chart.title = _chart_title_with_unit("핵심 재무비율 3개년 추이", "(단위: %)")
    chart.style = 10
    chart.y_axis.numFmt = "#,##0.0"
    _configure_axes(chart)

    for name in MAIN_RATIO_CHART_ITEMS:
        if name not in ratio_df.index:
            continue
        row = table_start + 1 + ratio_df.index.get_loc(name)
        values = Reference(ws, min_col=min_col, max_col=max_col, min_row=row, max_row=row)
        series = Series(values, title=name)
        series.smooth = False
        series.marker.symbol = "circle"
        chart.series.append(series)

    cats = Reference(ws, min_col=min_col, max_col=max_col, min_row=table_start, max_row=table_start)
    chart.set_categories(cats)

    ws.add_chart(chart, anchor_cell)


def _add_stacked_item_chart(
    ws: Worksheet,
    item_df: pd.DataFrame,
    table_start: int,
    anchor_cell: str,
    title_text: str,
    items: list[str],
) -> None:
    """비율이 동일해도 절대 규모/구성비가 달라질 수 있음을 보여주는 스택 바 차트.

    예: 유동비율이 3개년 내내 150%로 동일해도 유동자산 비중이 계속 줄고 비유동자산
    비중이 계속 느는 구조 변화는 비율 지표만으로는 드러나지 않는다. 두 항목을
    스택으로 쌓아 합계(자산총계 또는 부채총계) 대비 구성비 변화를 한눈에 보여준다.
    """
    period_cols = list(item_df.columns)
    if not period_cols:
        return

    min_col, max_col = 2, 1 + len(period_cols)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = _chart_title_with_unit(title_text, "(단위: 억원)")
    chart.style = 10
    chart.y_axis.numFmt = "#,##0"
    _configure_axes(chart)

    # from_rows=True 방식은 행이 1개뿐일 때 엑셀에서 계열/카테고리가 뒤바뀌어 보이는
    # 문제가 있어(실측: 유동비율 1개 지표 그래프에서 계열명이 "전전기/전기/당기"로
    # 잘못 표시됨), 항목마다 Series를 명시적으로 구성하는 방식으로 대체한다.
    for name in items:
        if name not in item_df.index:
            continue
        row = table_start + 1 + item_df.index.get_loc(name)
        values = Reference(ws, min_col=min_col, max_col=max_col, min_row=row, max_row=row)
        chart.series.append(Series(values, title=name))

    cats = Reference(ws, min_col=min_col, max_col=max_col, min_row=table_start, max_row=table_start)
    chart.set_categories(cats)

    ws.add_chart(chart, anchor_cell)


def _add_financials_sheet(wb: Workbook, wide_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("재무제표")
    ws.cell(row=1, column=1, value="3개년 재무제표 (단위: 원)").font = TITLE_FONT
    _write_df(ws, wide_df, 3, "계정과목")
    _autofit(ws, len(wide_df.columns) + 1)
    ws.freeze_panes = ws.cell(row=4, column=2)


def _add_info_sheet(
    wb: Workbook,
    info: dict,
    fs_div: str,
    missing_accounts: list[str],
    derived_notes: list[str],
    fs_note: str = "",
    financial_markers: list[str] | None = None,
) -> None:
    ws = wb.create_sheet("안내", 0)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    # fs_div는 계정과목 내용(예: 비지배지분 유무)으로 추론한 값이 아니라, 링크의 dcmNo가
    # 가리키는 첨부문서명 또는 DART API 응답 유무로 결정된 값이다(main.py 참고).
    # fs_note에는 그 판단 근거(자동판별/링크 기반 선택/폴백 여부)가 담겨 있다.
    fs_desc = "연결재무제표(CFS)" if fs_div == "CFS" else "개별/별도재무제표(OFS)"
    if fs_note:
        fs_desc += f" — {fs_note}"

    rows = [
        ("회사명", f"{info['corp_name']} (corp_code={info['corp_code']})"),
        ("공시서류명", info.get("report_nm", "")),
        ("사업연도(추정)", info.get("bsns_year", "")),
        ("재무제표 기준", fs_desc),
    ]
    r = 1
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        r += 1

    r += 1
    if financial_markers:
        warn_cell = ws.cell(
            row=r,
            column=1,
            value=(
                "⚠️ 이 회사는 금융업(은행/보험/지주 등)으로 추정됩니다. 본 도구의 "
                "표준 재무비율(유동비율 등)은 일반 기업 기준으로 설계되어 있어 "
                "다수 지표가 공란 처리될 수 있습니다. 금융업종은 예대율/BIS비율 등 "
                "별도 지표가 필요하며, 본 도구는 현재 이를 지원하지 않습니다."
            ),
        )
        warn_cell.font = Font(bold=True, color="C00000", size=12)
        warn_cell.fill = WARN_FILL
        r += 1
        ws.cell(row=r, column=1, value=f"(감지된 금융업 특유 계정: {', '.join(financial_markers)})")
        r += 2

    if missing_accounts:
        ws.cell(
            row=r,
            column=1,
            value=(
                "[경고] 다음 표준계정을 하나 이상의 기간에서 찾지 못해 관련 비율이 "
                "공란 처리되었습니다 (임의 추정하지 않음)"
            ),
        ).font = Font(bold=True, color="C00000")
        r += 1
        ws.cell(row=r, column=1, value=", ".join(missing_accounts))
        r += 1

    if derived_notes:
        r += 1
        ws.cell(row=r, column=1, value="[안내] 파생 계산된 항목").font = Font(bold=True)
        r += 1
        for note in derived_notes:
            ws.cell(row=r, column=1, value=note)
            r += 1

    r += 1
    ws.cell(
        row=r,
        column=1,
        value="[면책] 본 결과는 참고용 사전분석 자료이며 감사의견 형성의 유일한 근거가 될 수 없습니다.",
    ).font = Font(italic=True)


def _add_bs_detail_sheet(wb: Workbook, bs_item_df: pd.DataFrame) -> None:
    """"상세보기 — 원 항목 추이" 시트.

    비율 그래프(메인, "재무비율" 시트)는 정보 과부하를 피하기 위해 여기에는 넣지 않고,
    비율만으로는 드러나지 않는 절대금액 규모/구성비 변화를 보여주는 원 항목 스택 차트만
    별도 탭으로 분리한다. 이상 신호를 비율 그래프에서 먼저 포착한 뒤, 필요할 때만 이
    시트로 넘어와 원인을 파고드는 감사 위험평가 워크플로우에 대응한다.
    """
    ws = wb.create_sheet("상세보기_원항목추이")
    ws.cell(row=1, column=1, value="상세보기 — 원 항목(절대금액) 추이").font = TITLE_FONT
    ws.cell(
        row=2,
        column=1,
        value=(
            "[안내] 비율(예: 유동비율)이 3개년 내내 동일해도 아래처럼 절대 규모나 "
            "구성비가 크게 바뀌었다면 사업 축소·자산 재평가 등 추가 확인이 필요할 수 있습니다."
        ),
    ).font = Font(italic=True)

    table_start = _add_bs_item_table(ws, bs_item_df, title_row=4)
    n_cols = len(bs_item_df.columns) + 1
    chart_anchor_col = get_column_letter(n_cols + 2)

    _add_stacked_item_chart(
        ws,
        bs_item_df,
        table_start,
        f"{chart_anchor_col}{table_start}",
        "자산 구성 추이 (유동자산 vs 비유동자산)",
        ASSET_STRUCTURE_ITEMS,
    )
    _add_stacked_item_chart(
        ws,
        bs_item_df,
        table_start,
        f"{chart_anchor_col}{table_start + 30}",
        "부채 구성 추이 (유동부채 vs 비유동부채)",
        LIABILITY_STRUCTURE_ITEMS,
    )


def _add_screening_sheet(
    wb: Workbook,
    screening_df: pd.DataFrame,
    excluded_notes: list[str],
    threshold_pct: float,
    strong_threshold_pct: float,
    trend_deviation_threshold_pctp: float,
) -> None:
    """"스크리닝_전기당기증감" 시트 — 감사 위험평가용 분석적 절차(analytical procedures).

    표준 8개 비율과 달리 재무상태표/손익계산서/현금흐름표에 속한 회사의 모든
    개별 계정(수십~수백 개)을 대상으로, 전기 대비 당기 증감이 큰 계정을 찾아낸다.
    """
    ws = wb.create_sheet("스크리닝_전기당기증감")
    ws.cell(
        row=1, column=1, value="전기 대비 당기 증감 스크리닝 (분석적 절차)"
    ).font = TITLE_FONT
    ws.cell(
        row=2,
        column=1,
        value=(
            f"[안내] 재무상태표/손익계산서/현금흐름표의 개별 계정 중 전기 대비 증감률이 "
            f"±{threshold_pct:.0f}%를 초과하는 계정만 추립니다(중요성 기준 미만 소액 계정은 "
            f"제외, 포함 여부는 전기→당기 기준). 신규계정(전기=0)·부호전환(흑자↔적자)은 %를 "
            f"계산하지 않고 별도 표시합니다. 재무제표구분(재무상태표→손익계산서→현금흐름표) "
            f"순으로 묶은 뒤 그 안에서 증감액(절대금액) 내림차순 정렬하며, "
            f"±{strong_threshold_pct:.0f}% 초과인 계정은 '전기→당기 증감률(%)' 칸이 노란색으로 "
            f"강조됩니다. '구분'이 \"증감률초과\"인 항목은 통상적으로 계산된 증감률이 기준을 "
            f"넘었다는 뜻이며(문제가 없다는 뜻이 아님), 신규계정/부호전환은 애초에 %를 계산할 "
            f"수 없어 별도 분류된 것입니다. '전전기→전기 증감률(%)'은 비교 참고용이며, 두 "
            f"구간의 증감률 차이가 {trend_deviation_threshold_pctp:.0f}%p 이상이면(둘 다 정상 "
            f"계산된 경우에 한함) '추세이탈'에 표시됩니다 — 단순히 증감률이 큰 게 아니라 "
            f"이번 기에 갑자기 패턴이 바뀐 계정을 잡아내기 위함입니다."
        ),
    ).font = Font(italic=True)

    if excluded_notes:
        r = 4
        ws.cell(
            row=r,
            column=1,
            value="[경고] 동일 계정명으로 매칭된 금액이 서로 달라 스크리닝에서 제외된 항목",
        ).font = Font(bold=True, color="C00000")
        r += 1
        for note in excluded_notes:
            ws.cell(row=r, column=1, value=note)
            r += 1
        table_start = r + 1
    else:
        table_start = 4

    header_row = table_start
    for j, col in enumerate(screening.SCREENING_COLUMNS, start=1):
        ws.cell(row=header_row, column=j, value=col)
    _style_header_row(ws, header_row, len(screening.SCREENING_COLUMNS))

    col_idx = {name: i + 1 for i, name in enumerate(screening.SCREENING_COLUMNS)}
    amount_cols = ("전전기금액", "전기금액", "당기금액", "증감액")
    pct_cols = ("전기→당기 증감률(%)", "전전기→전기 증감률(%)")

    r = header_row + 1
    for _, row in screening_df.iterrows():
        ws.cell(row=r, column=col_idx["재무제표구분"], value=row["재무제표구분"])
        ws.cell(row=r, column=col_idx["계정명"], value=row["계정명"])
        for key in amount_cols:
            value = row[key]
            if not (isinstance(value, float) and math.isnan(value)):
                cell = ws.cell(row=r, column=col_idx[key], value=round(float(value), 2))
                cell.number_format = "#,##0"
        for key in pct_cols:
            value = row[key]
            if not (isinstance(value, float) and math.isnan(value)):
                cell = ws.cell(row=r, column=col_idx[key], value=round(float(value), 2))
                cell.number_format = "#,##0.00"
        ws.cell(row=r, column=col_idx["추세이탈"], value=row["추세이탈"] or None)
        ws.cell(row=r, column=col_idx["구분"], value=row["구분"])
        r += 1
    last_row = r - 1

    if not screening_df.empty:
        pct_col_letter = get_column_letter(col_idx["전기→당기 증감률(%)"])
        cell_range = f"{pct_col_letter}{header_row + 1}:{pct_col_letter}{last_row}"
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="greaterThan",
                formula=[str(strong_threshold_pct)],
                fill=SCREENING_HIGHLIGHT_FILL,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="lessThan",
                formula=[str(-strong_threshold_pct)],
                fill=SCREENING_HIGHLIGHT_FILL,
            ),
        )
    else:
        ws.cell(row=header_row + 1, column=1, value="(중요성·임계값 기준을 초과하는 계정 없음)")

    _autofit(ws, len(screening.SCREENING_COLUMNS))
    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)


def _write_check_table(
    ws: Worksheet,
    title: str,
    headers: list[str],
    checks: list[verification.CheckResult],
    start_row: int,
) -> int:
    """검증 결과 목록을 표로 쓴다. 반환값은 다음 콘텐츠를 이어 쓸 행 번호."""
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True)
    header_row = start_row + 1
    for j, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=j, value=header)
    _style_header_row(ws, header_row, len(headers))

    r = header_row + 1
    for check in checks:
        ws.cell(row=r, column=1, value=check.label)
        ws.cell(row=r, column=2, value=check.period)
        expected_cell = ws.cell(row=r, column=3, value=round(check.expected, 2))
        expected_cell.number_format = "#,##0.00"
        actual_cell = ws.cell(row=r, column=4, value=round(check.actual, 2))
        actual_cell.number_format = "#,##0.00"
        ws.cell(row=r, column=5, value="일치" if check.ok else "불일치")
        if not check.ok:
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = WARN_FILL
        r += 1

    if not checks:
        ws.cell(row=r, column=1, value="(검증 대상 없음 — 관련 계정 매핑 실패로 이미 별도 경고 표시됨)")
        r += 1

    return r


def _add_verification_sheet(
    wb: Workbook,
    ratio_checks: list[verification.CheckResult],
    bs_checks: list[verification.CheckResult],
    screening_checks: list[verification.CheckResult] | None = None,
) -> None:
    """"검증" 시트 — ratios.py 계산 로직과 완전히 별개의 코드로 재계산한 대사 결과.

    회계사가 산출물을 받을 때마다 수작업으로 돌리던 대사 절차(비율 재계산 비교,
    자산=부채+자본 등식 확인)를 엑셀 생성 파이프라인 자체에 내장했다. 모든 항목이
    "일치"로 나와야 정상이며, "불일치"가 하나라도 있으면 계산 로직 또는 원본
    데이터에 문제가 있다는 뜻이다.
    """
    ws = wb.create_sheet("검증")
    ws.cell(row=1, column=1, value="독립 검증 (재계산 대사 + 재무상태표 등식)").font = TITLE_FONT
    ws.cell(
        row=2,
        column=1,
        value=(
            "[안내] 아래 값은 재무비율/원 항목 시트를 만든 계산 로직과 별개의 코드로 "
            "재계산한 결과입니다. 모두 '일치'가 정상이며, '불일치'가 있으면 계산 로직이나 "
            "원본 데이터에 문제가 있다는 뜻이니 반드시 원인을 확인하세요."
        ),
    ).font = Font(italic=True)

    screening_checks = screening_checks or []
    all_checks = ratio_checks + bs_checks + screening_checks
    failed = [c for c in all_checks if not c.ok]
    if failed:
        summary_cell = ws.cell(
            row=4,
            column=1,
            value=f"⚠ 불일치 {len(failed)}건 발견 (전체 {len(all_checks)}건 중) — 아래 표에서 확인하세요",
        )
        summary_cell.font = Font(bold=True, color="C00000")
    else:
        summary_cell = ws.cell(
            row=4, column=1, value=f"✓ 전체 {len(all_checks)}건 모두 일치 (검증 통과)"
        )
        summary_cell.font = Font(bold=True, color="006100")

    row = 6
    row = _write_check_table(
        ws,
        "1) 핵심 비율 재계산 대사 (유동비율/부채비율/자기자본비율)",
        ["비율명", "기간", "보고값(%)", "재계산값(%)", "일치여부"],
        ratio_checks,
        row,
    )
    row += 2
    row = _write_check_table(
        ws,
        "2) 재무상태표 등식 검증 (단위: 원)",
        ["항목", "기간", "좌변(합계)", "우변(구성요소 합)", "일치여부"],
        bs_checks,
        row,
    )
    row += 2
    _write_check_table(
        ws,
        "3) 전기 대비 당기 증감 스크리닝 검증 (원본금액 대사 + 증감률/분류 재계산 + 전수 검사)",
        ["검증항목", "기간", "기준값", "재계산값", "일치여부"],
        screening_checks,
        row,
    )

    _autofit(ws, 5)


def build_workbook(
    info: dict,
    fs_div: str,
    wide_df: pd.DataFrame,
    ratio_df: pd.DataFrame,
    missing_accounts: list[str],
    derived_notes: list[str],
    bs_item_df: pd.DataFrame,
    bs_missing_items: list[str],
    accounts_wide: pd.DataFrame,
    fs_note: str = "",
    financial_markers: list[str] | None = None,
    blank_reasons: dict[str, dict[str, str]] | None = None,
    long_df: pd.DataFrame | None = None,
    screening_df: pd.DataFrame | None = None,
    screening_excluded_notes: list[str] | None = None,
    screening_threshold_pct: float = 10.0,
    screening_strong_threshold_pct: float = 30.0,
    screening_materiality_pct: float = 1.0,
    screening_trend_deviation_threshold_pctp: float = screening.TREND_DEVIATION_THRESHOLD_PCTP,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    combined_missing = missing_accounts + [
        name for name in bs_missing_items if name not in missing_accounts
    ]
    _add_info_sheet(
        wb,
        info,
        fs_div,
        combined_missing,
        derived_notes,
        fs_note=fs_note,
        financial_markers=financial_markers,
    )
    _add_financials_sheet(wb, wide_df)
    _add_ratio_sheet(wb, ratio_df, blank_reasons=blank_reasons)
    _add_bs_detail_sheet(wb, bs_item_df)

    screening_checks: list[verification.CheckResult] = []
    if screening_df is not None:
        _add_screening_sheet(
            wb,
            screening_df,
            screening_excluded_notes or [],
            screening_threshold_pct,
            screening_strong_threshold_pct,
            screening_trend_deviation_threshold_pctp,
        )
        if long_df is not None:
            total_assets = (
                accounts_wide.loc["자산총계", "당기"]
                if "자산총계" in accounts_wide.index and "당기" in accounts_wide.columns
                else math.nan
            )
            screening_checks = verification.verify_screening(
                long_df,
                screening_df,
                total_assets,
                threshold_pct=screening_threshold_pct,
                materiality_pct=screening_materiality_pct,
                trend_deviation_threshold_pctp=screening_trend_deviation_threshold_pctp,
            )

    ratio_checks = verification.verify_ratios(accounts_wide, ratio_df)
    bs_checks = verification.verify_balance_sheet_identity(accounts_wide)
    _add_verification_sheet(wb, ratio_checks, bs_checks, screening_checks)

    wb.active = 0
    return wb
